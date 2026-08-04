"""
Main CT-QC extraction engine.
"""

from __future__ import annotations

import hashlib
import logging
import math
import re
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl as pyx
import pandas as pd
from openpyxl.utils import column_index_from_string, range_boundaries
from openpyxl.utils.datetime import from_excel

from core.mapping_parser import TemplateMappingParser
from models.extraction_result import ExtractionResult
from models.mapping import ScalarSpec, SectionSpec, SheetSpec, TableSpec, TemplateSpec

from .normalizers import Normalizers
from .unit_converter import UnitConverter
from .validators import Validators

logger = logging.getLogger(__name__)


class ExtractionEngine:
    """
    Extracts CT-QC data from Excel workbooks using compiled YAML mappings.
    """

    def __init__(self, config: Dict):
        self.config = config
        self.mapping_parser = TemplateMappingParser()
        self.normalizers = Normalizers(config)
        self.validators = Validators(config)
        self.unit_converter = UnitConverter(config)

    def extract_from_path(self, excel_path: str, sidecar: Dict[str, Any]) -> ExtractionResult:
        logger.info(f"Extraction des donnees depuis: {excel_path}")

        wb = None
        try:
            wb = pyx.load_workbook(excel_path, read_only=True, data_only=True)
            return self.extract_from_workbook(wb, sidecar, excel_path)
        except Exception as e:
            logger.error(f"Echec de l'extraction pour {excel_path}: {e}")
            raise
        finally:
            if wb:
                wb.close()

    def extract_from_workbook(
        self, wb, sidecar: Dict[str, Any], source_path: str = ""
    ) -> ExtractionResult:
        template_id = sidecar.get("id") or sidecar.get("_template_id") or "ctqc_base"
        template = self.mapping_parser.compile(sidecar, template_id)
        meta = self._collect_metadata(wb, template, source_path)

        fields: Dict[str, Any] = {}
        tables: Dict[str, pd.DataFrame] = {}
        scalar_records: List[Dict[str, Any]] = []
        named_records: List[Dict[str, Any]] = []
        audit = {
            "sheets_processed": [],
            "errors": [],
            "warnings": [],
            "stats": {
                "total_sheets": len(wb.sheetnames),
                "configured_sheets": len(template.sheets),
                "processed_sheets": 0,
                "extracted_fields": 0,
                "extracted_tables": 0,
                "extracted_scalar_records": 0,
                "extracted_named_results": 0,
            },
        }

        workbook_sheets = set(wb.sheetnames)
        for sheet_spec in template.sheets.values():
            sheet_audit = {
                "sheet": sheet_spec.sheet_name,
                "sheet_id": sheet_spec.sheet_id,
                "fields_extracted": 0,
                "tables_extracted": 0,
                "scalar_records": 0,
                "named_results": 0,
                "errors": [],
                "warnings": [],
                "skipped": False,
            }

            if not sheet_spec.extract:
                sheet_audit["skipped"] = True
                audit["sheets_processed"].append(sheet_audit)
                continue

            if sheet_spec.sheet_name not in workbook_sheets:
                message = f"Configured sheet missing: {sheet_spec.sheet_name}"
                if sheet_spec.required:
                    sheet_audit["errors"].append(message)
                    audit["errors"].append(message)
                else:
                    sheet_audit["warnings"].append(message)
                    audit["warnings"].append(message)
                audit["sheets_processed"].append(sheet_audit)
                continue

            try:
                sheet_result = self._process_sheet(wb[sheet_spec.sheet_name], sheet_spec)
                self._merge_fields(fields, sheet_result.fields)
                self._merge_tables(tables, sheet_result.tables)

                if not sheet_result.scalar_records.empty:
                    scalar_records.extend(sheet_result.scalar_records.to_dict("records"))
                if not sheet_result.named_results.empty:
                    named_records.extend(sheet_result.named_results.to_dict("records"))

                sheet_audit["fields_extracted"] = len(sheet_result.fields)
                sheet_audit["tables_extracted"] = len(sheet_result.tables)
                sheet_audit["scalar_records"] = len(sheet_result.scalar_records)
                sheet_audit["named_results"] = len(sheet_result.named_results)
                sheet_audit["errors"] = sheet_result.audit.get("errors", [])
                sheet_audit["warnings"] = sheet_result.audit.get("warnings", [])

                audit["stats"]["processed_sheets"] += 1
                audit["stats"]["extracted_fields"] += len(sheet_result.fields)
                audit["stats"]["extracted_tables"] += len(sheet_result.tables)
                audit["stats"]["extracted_scalar_records"] += len(sheet_result.scalar_records)
                audit["stats"]["extracted_named_results"] += len(sheet_result.named_results)
            except Exception as e:
                message = f"Erreur traitement feuille {sheet_spec.sheet_name}: {e}"
                sheet_audit["errors"].append(message)
                audit["errors"].append(message)
                logger.error(message)

            audit["sheets_processed"].append(sheet_audit)

        logger.info(
            "Extraction terminee: %s champs, %s tables, %s named results",
            audit["stats"]["extracted_fields"],
            audit["stats"]["extracted_tables"],
            audit["stats"]["extracted_named_results"],
        )

        return ExtractionResult(
            fields=fields,
            tables=tables,
            audit=audit,
            meta=meta,
            scalar_records=pd.DataFrame(scalar_records),
            named_results=pd.DataFrame(named_records),
        )

    def _process_sheet(self, sheet, sheet_spec: SheetSpec) -> ExtractionResult:
        fields: Dict[str, Any] = {}
        tables: Dict[str, pd.DataFrame] = {}
        scalar_records: List[Dict[str, Any]] = []
        named_records: List[Dict[str, Any]] = []
        audit = {"errors": [], "warnings": []}

        for section in sheet_spec.sections.values():
            try:
                section_metadata = self._extract_scalar_group(
                    sheet, sheet_spec, section, section.metadata, scalar_records, audit
                )

                scalar_values = self._extract_scalar_group(
                    sheet, sheet_spec, section, section.scalars, scalar_records, audit
                )
                self._merge_fields(fields, scalar_values)

                self._extract_scalar_group(
                    sheet, sheet_spec, section, section.result_scalars, scalar_records, audit
                )

                named_values = self._extract_scalar_group(
                    sheet, sheet_spec, section, section.named_results, named_records, audit
                )
                if named_values:
                    self._merge_fields(fields, {f"{section.section_id}.{k}": v for k, v in named_values.items()})

                for table_spec in section.tables:
                    df = self._extract_table(sheet, sheet_spec, section, table_spec, section_metadata)
                    if df is None or df.empty:
                        continue

                    table_key = table_spec.db_table or section.db_table or table_spec.name
                    self._merge_tables(tables, {table_key: df})
            except Exception as e:
                message = f"Section {section.section_id}: {e}"
                audit["errors"].append(message)
                logger.error(message)

        return ExtractionResult(
            fields=fields,
            tables=tables,
            audit=audit,
            meta={},
            scalar_records=pd.DataFrame(scalar_records),
            named_results=pd.DataFrame(named_records),
        )

    def _extract_scalar_group(
        self,
        sheet,
        sheet_spec: SheetSpec,
        section: SectionSpec,
        scalars: Dict[str, ScalarSpec],
        records: List[Dict[str, Any]],
        audit: Dict[str, List[str]],
    ) -> Dict[str, Any]:
        values: Dict[str, Any] = {}
        for scalar in scalars.values():
            try:
                value = self._extract_scalar_value(sheet, scalar.config)
                if value is None and self._is_required(scalar.config):
                    audit["errors"].append(f"Champ requis manquant: {section.section_id}.{scalar.name}")
                values[scalar.name] = value
                records.append(
                    self._scalar_record(sheet_spec, section, scalar, value, sheet_spec.inject_columns)
                )
            except Exception as e:
                audit["errors"].append(f"Champ {section.section_id}.{scalar.name}: {e}")
                logger.error(f"Erreur champ {section.section_id}.{scalar.name}: {e}")
        return values

    def _extract_scalar_value(self, sheet, config: Dict[str, Any]) -> Any:
        dtype = self._dtype(config)
        if dtype == "list_of_dict":
            return self._extract_list_of_dict(sheet, config)
        if dtype == "list_of_string":
            return self._extract_list_of_string(sheet, config)

        value = None
        if "row" in config and "col" in config:
            value = self._cell_value(sheet, config["row"], config["col"])
        elif "locator" in config:
            locator = config.get("locator") or {}
            if "cell" in locator:
                value = self._extract_cell_ref(sheet, locator["cell"])
            elif "find" in locator:
                value = self._extract_by_pattern(sheet, locator["find"])
        elif "name" in config:
            value = self._extract_defined_name(sheet, config["name"])

        if value is None:
            value = config.get("default")

        return self._convert_value(value, config)

    def _extract_list_of_dict(self, sheet, config: Dict[str, Any]) -> List[Dict[str, Any]]:
        start = int(config.get("data_start_row") or 1)
        end = int(config.get("data_end_row") or config.get("end_row") or start)
        col_nom = config.get("col_nom") or config.get("name_col") or config.get("col")
        col_riziv = config.get("col_riziv") or config.get("riziv_col")
        rows = []

        for row in range(start, end + 1):
            first_value = self._cell_value(sheet, row, col_nom) if col_nom else None
            if self._is_empty(first_value):
                break
            item = {"nom": self._convert_value(first_value, {"dtype": "string"})}
            if col_riziv:
                item["riziv"] = self._convert_value(
                    self._cell_value(sheet, row, col_riziv), {"dtype": "string"}
                )
            rows.append(item)
        return rows

    def _extract_list_of_string(self, sheet, config: Dict[str, Any]) -> List[str]:
        start = int(config.get("data_start_row") or 1)
        end = int(config.get("data_end_row") or config.get("end_row") or start)
        col = config.get("col")
        values = []
        for row in range(start, end + 1):
            value = self._cell_value(sheet, row, col)
            if self._is_empty(value):
                break
            values.append(self._convert_value(value, {"dtype": "string"}))
        return values

    def _extract_table(
        self,
        sheet,
        sheet_spec: SheetSpec,
        section: SectionSpec,
        table_spec: TableSpec,
        section_metadata: Dict[str, Any],
    ) -> Optional[pd.DataFrame]:
        config = table_spec.config

        if "locator" in config and "range" in config["locator"]:
            df = self._extract_legacy_range(sheet, config["locator"]["range"], table_spec)
        else:
            rows = self._resolve_table_rows(sheet, table_spec)
            data = []
            for row in rows:
                if self._should_skip_row(sheet, row, config):
                    continue

                record: Dict[str, Any] = {}
                has_data = False
                for column in table_spec.columns.values():
                    source_row = int(column.config.get("row", row))
                    raw_value = self._cell_value(sheet, source_row, column.config.get("col"))
                    if source_row == row and not self._is_empty(raw_value):
                        has_data = True
                    record[column.name] = self._convert_value(raw_value, column.config)

                if not has_data and table_spec.columns:
                    continue
                data.append(record)

            df = pd.DataFrame(data)

        if df is None or df.empty:
            return None

        for key, value in section_metadata.items():
            if key not in df.columns:
                df[key] = value

        self._apply_inject_columns(df, sheet_spec.inject_columns)
        return df

    def _extract_legacy_range(
        self, sheet, range_str: str, table_spec: TableSpec
    ) -> Optional[pd.DataFrame]:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(range_str)
        except Exception as e:
            logger.error(f"Erreur extraction plage {range_str}: {e}")
            return None

        headers = []
        for col in range(min_col, max_col + 1):
            value = sheet.cell(row=min_row, column=col).value
            headers.append(str(value).strip() if value is not None else f"Column_{col}")

        data = []
        for row in range(min_row + 1, max_row + 1):
            row_values = []
            has_data = False
            for col in range(min_col, max_col + 1):
                value = sheet.cell(row=row, column=col).value
                if not self._is_empty(value):
                    has_data = True
                row_values.append(value)
            if has_data:
                data.append(row_values)

        if not data:
            return None

        df = pd.DataFrame(data, columns=headers).dropna(axis=1, how="all")
        configured_columns = list(table_spec.columns.keys())
        if configured_columns:
            if len(df.columns) >= len(configured_columns):
                mapping = {df.columns[i]: configured_columns[i] for i in range(len(configured_columns))}
                df = df.rename(columns=mapping)[configured_columns]
            else:
                for i, col_name in enumerate(configured_columns):
                    if i < len(df.columns):
                        df = df.rename(columns={df.columns[i]: col_name})
                    else:
                        df[col_name] = None

            for column in table_spec.columns.values():
                if column.name in df.columns:
                    df[column.name] = df[column.name].apply(
                        lambda value, cfg=column.config: self._convert_value(value, cfg)
                    )
        return df

    def _resolve_table_rows(self, sheet, table_spec: TableSpec) -> List[int]:
        config = table_spec.config
        start = int(config.get("data_start_row") or config.get("start_row") or 1)
        explicit_end = config.get("end_row") or config.get("data_end_row")
        if explicit_end is not None:
            return list(range(start, int(explicit_end) + 1))

        condition = config.get("end_condition") or {}
        condition_type = condition.get("type")
        if condition_type == "n_rows":
            return list(range(start, start + int(condition.get("n", 0))))

        max_rows = min(
            int(self.config.get("engine", {}).get("extraction", {}).get("max_rows_per_sheet", 10000)),
            sheet.max_row or 10000,
        )

        if condition_type in {"empty_col", "empty_cell"}:
            col = condition.get("col") or self._first_table_col(table_spec)
            return self._rows_until_empty(sheet, start, max_rows, col)

        if condition_type == "label_row":
            col = condition.get("col") or self._first_table_col(table_spec)
            label = self._normalize_text(condition.get("label"))
            rows = []
            for row in range(start, max_rows + 1):
                value = self._normalize_text(self._cell_value(sheet, row, col))
                if value == label:
                    break
                rows.append(row)
            return rows

        col = self._first_table_col(table_spec)
        return self._rows_until_empty(sheet, start, max_rows, col)

    def _rows_until_empty(self, sheet, start: int, max_rows: int, col: Any) -> List[int]:
        rows = []
        for row in range(start, max_rows + 1):
            if self._is_empty(self._cell_value(sheet, row, col)):
                break
            rows.append(row)
        return rows

    def _first_table_col(self, table_spec: TableSpec) -> Any:
        for column in table_spec.columns.values():
            if "col" in column.config:
                return column.config["col"]
        return "A"

    def _should_skip_row(self, sheet, row: int, config: Dict[str, Any]) -> bool:
        if "skip_if_value" not in config or "skip_col" not in config:
            return False
        actual = self._normalize_text(self._cell_value(sheet, row, config["skip_col"]))
        expected = config["skip_if_value"]
        if isinstance(expected, list):
            return actual in {self._normalize_text(value) for value in expected}
        return actual == self._normalize_text(expected)

    def _apply_inject_columns(self, df: pd.DataFrame, inject_columns: Dict[str, Any]) -> None:
        for key, value in (inject_columns or {}).items():
            if key in df.columns:
                logger.warning(f"Injected column already exists and will be left unchanged: {key}")
                continue
            df[key] = value

    def _scalar_record(
        self,
        sheet_spec: SheetSpec,
        section: SectionSpec,
        scalar: ScalarSpec,
        value: Any,
        inject_columns: Dict[str, Any],
    ) -> Dict[str, Any]:
        record = {
            "sheet_id": sheet_spec.sheet_id,
            "sheet_name": sheet_spec.sheet_name,
            "section_id": section.section_id,
            "db_table": section.db_table,
            "name": scalar.name,
            "kind": scalar.kind,
            "dtype": self._dtype(scalar.config),
            "row": scalar.config.get("row"),
            "col": scalar.config.get("col"),
            "value": value,
        }
        record.update(inject_columns or {})
        return record

    def _convert_value(self, value: Any, config: Dict[str, Any]) -> Any:
        dtype = self._dtype(config)
        if value is None:
            return None
        if isinstance(value, float) and math.isnan(value):
            return None
        if isinstance(value, str) and value.strip() == "":
            return None

        if "normalize" in config:
            value = self.normalizers.apply(value, config["normalize"])

        try:
            if dtype in {"string", "str"}:
                return str(value).strip()
            if dtype in {"float", "number"}:
                return self._to_float(value, config)
            if dtype == "int":
                number = self._to_float(value, config)
                return int(number) if number is not None else None
            if dtype in {"bool", "boolean"}:
                return self._to_bool(value, config)
            if dtype == "date":
                converted = self._to_datetime(value)
                return converted.date().isoformat() if isinstance(converted, datetime) else converted
            if dtype == "datetime":
                converted = self._to_datetime(value)
                return converted.isoformat() if isinstance(converted, datetime) else converted
            return value
        except Exception as e:
            logger.debug(f"Conversion type echouee: {value} -> {dtype}, erreur: {e}")
            return None

    def _to_float(self, value: Any, config: Dict[str, Any]) -> Optional[float]:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return float(value)
        text = str(value).strip().replace("\u00a0", " ")
        if self._normalize_text(text) in {"nvt", "ngemeten", "na", "n/a"}:
            return None
        text = text.replace("%", "").strip()
        if config.get("locale_safe") or config.get("localesafe") or "," in text:
            text = text.replace(" ", "")
            if "," in text and "." not in text:
                text = text.replace(",", ".")
        return float(text)

    def _to_bool(self, value: Any, config: Dict[str, Any]) -> Optional[bool]:
        if isinstance(value, bool):
            return value
        bool_map = config.get("bool_map") or config.get("boolmap") or {}
        if bool_map:
            normalized_map = {self._normalize_text(k): bool(v) for k, v in bool_map.items()}
            normalized = self._normalize_text(value)
            if normalized in normalized_map:
                return normalized_map[normalized]
        normalized = self._normalize_text(value)
        if normalized in {"true", "1", "yes", "y", "ja", "oui", "vrai"}:
            return True
        if normalized in {"false", "0", "no", "n", "nee", "non", "faux"}:
            return False
        return None

    def _to_datetime(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value
        if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
            return datetime(value.year, value.month, value.day)
        if isinstance(value, (int, float)):
            return from_excel(value)
        parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
        if pd.isna(parsed):
            return value
        return parsed.to_pydatetime()

    def _dtype(self, config: Dict[str, Any]) -> str:
        return str(config.get("dtype") or config.get("type") or "string").lower()

    def _is_required(self, config: Dict[str, Any]) -> bool:
        if config.get("optional") is True:
            return False
        return bool(config.get("required", False))

    def _cell_value(self, sheet, row: Any, col: Any) -> Any:
        if row is None or col is None:
            return None
        return sheet.cell(row=int(row), column=self._col_index(col)).value

    def _extract_cell_ref(self, sheet, cell_ref: str) -> Any:
        try:
            return sheet[cell_ref].value
        except Exception:
            return None

    def _extract_defined_name(self, sheet, defined_name: str) -> Any:
        try:
            if re.match(r"^[A-Z]+\d+$", defined_name):
                return self._extract_cell_ref(sheet, defined_name)

            wb = sheet.parent
            if defined_name in wb.defined_names:
                defined_obj = wb.defined_names[defined_name]
                cell_ref = defined_obj.value
                if "!" in cell_ref:
                    sheet_name, cell_addr = cell_ref.split("!")
                    target_sheet = wb[sheet_name.strip("'")]
                    return self._extract_cell_ref(target_sheet, cell_addr)

            return self._find_value_by_text(sheet, defined_name)
        except Exception as e:
            logger.debug(f"Echec extraction nom defini {defined_name}: {e}")
            return None

    def _find_value_by_text(self, sheet, search_text: str) -> Any:
        max_rows = min(100, sheet.max_row or 100)
        max_cols = min(50, sheet.max_column or 50)
        for row in range(1, max_rows + 1):
            for col in range(1, max_cols + 1):
                value = sheet.cell(row=row, column=col).value
                if value and search_text in str(value):
                    return sheet.cell(row=row, column=col + 1).value
        return None

    def _extract_by_pattern(self, sheet, pattern_config: Dict[str, Any]) -> Any:
        search_text = pattern_config.get("text", "")
        start_row = int(pattern_config.get("start_row", 1))
        end_row = int(pattern_config.get("end_row", 100))
        offset_col = int(pattern_config.get("offset_col", 1))

        for row in range(start_row, end_row + 1):
            for col in range(1, 50):
                value = sheet.cell(row=row, column=col).value
                if value and search_text in str(value):
                    return sheet.cell(row=row, column=col + offset_col).value
        return None

    def _col_index(self, col: Any) -> int:
        if isinstance(col, int):
            return col
        if isinstance(col, str) and col.isdigit():
            return int(col)
        return column_index_from_string(str(col).strip())

    def _is_empty(self, value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, float) and math.isnan(value):
            return True
        return isinstance(value, str) and value.strip() == ""

    def _normalize_text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().lower()

    def _merge_fields(self, target: Dict[str, Any], source: Dict[str, Any]) -> None:
        for key, value in source.items():
            if key not in target:
                target[key] = value
            else:
                suffix = 2
                next_key = f"{key}_{suffix}"
                while next_key in target:
                    suffix += 1
                    next_key = f"{key}_{suffix}"
                target[next_key] = value

    def _merge_tables(self, target: Dict[str, pd.DataFrame], source: Dict[str, pd.DataFrame]) -> None:
        for key, df in source.items():
            if key in target:
                target[key] = pd.concat([target[key], df], ignore_index=True)
            else:
                target[key] = df

    def _collect_metadata(self, wb, template: TemplateSpec, source_path: str = "") -> Dict[str, Any]:
        sheets_content = "|".join(wb.sheetnames)
        source_filename = source_path.replace("\\", "/").split("/")[-1] if source_path else ""
        return {
            "template_id": template.template_id,
            "sidecar_version": template.version,
            "template_schema": template.schema,
            "workbook_sheets": wb.sheetnames,
            "workbook_hash": hashlib.sha256(sheets_content.encode()).hexdigest()[:16],
            "source_path": source_path,
            "source_filename": source_filename,
            "run_id": uuid.uuid4().hex,
            "extraction_timestamp": datetime.now().isoformat(),
            "engine_version": "3.0.0",
        }
