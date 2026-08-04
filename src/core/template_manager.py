"""
Template Manager - Gestion unifiée des templates Excel et sidecars YAML
"""

import os
import yaml
from typing import Dict, Any, Optional, List
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class TemplateManager:
    """
    Gère les templates Excel et leurs sidecars YAML associés
    """
    
    def __init__(self, templates_dir: str):
        self.templates_dir = Path(templates_dir)
        self.masters_dir = self.templates_dir / "masters"
        self.registry = self._load_registry()
        
    def _load_registry(self) -> Dict:
        """Charge le registry des templates"""
        registry_path = self.templates_dir / "index.yml"
        if not registry_path.exists():
            logger.warning(f"Registry not found: {registry_path}")
            return {"templates": {}}
        
        try:
            with open(registry_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        except Exception as e:
            logger.error(f"Failed to load registry: {e}")
            return {"templates": {}}
    
    def get_template_info(self, template_id: str) -> Dict[str, Any]:
        """Retourne les informations complètes d'un template"""
        template_config = self.registry.get("templates", {}).get(template_id, {})
        
        # Chemins des fichiers
        excel_path = self.masters_dir / f"{template_id}.xlsx"
        sidecar_path = self._get_template_sidecar_path(template_id)
        
        return {
            'id': template_id,
            'excel_path': excel_path if excel_path.exists() else None,
            'sidecar_path': sidecar_path if sidecar_path and sidecar_path.exists() else None,
            'config': template_config,
            'excel_exists': excel_path.exists(),
            'sidecar_exists': bool(sidecar_path and sidecar_path.exists()),
            'description': template_config.get('description', 'No description')
        }
    
    def get_available_templates(self) -> List[Dict[str, Any]]:
        """Retourne la liste de tous les templates disponibles"""
        templates = []
        
        for template_id in self.registry.get("templates", {}).keys():
            template_info = self.get_template_info(template_id)
            if template_info['excel_exists'] or template_info['sidecar_exists']:
                templates.append(template_info)
        
        return templates
    
    def validate_template(self, template_id: str) -> Dict[str, Any]:
        """Valide la cohérence d'un template"""
        info = self.get_template_info(template_id)
        issues = []
        
        if not info['excel_exists']:
            issues.append(f"Excel template missing: {template_id}.xlsx")
        
        if not info['sidecar_exists']:
            issues.append(f"Sidecar missing: {template_id}.yml")
        
        # Vérification de l'héritage
        if info['config'].get('extends'):
            parent_id = info['config']['extends']
            parent_info = self.get_template_info(parent_id)
            if not parent_info['excel_exists']:
                issues.append(f"Parent template missing: {parent_id}")
        
        return {
            'template_id': template_id,
            'valid': len(issues) == 0,
            'issues': issues,
            'info': info
        }
    
    def get_template_excel_path(self, template_id: str) -> Optional[Path]:
        """Retourne le chemin du template Excel"""
        # Essai avec différentes extensions
        for ext in ['.xlsx', '.xlsm', '.xltx']:
            excel_path = self.masters_dir / f"{template_id}{ext}"
            if excel_path.exists():
                return excel_path
        
        return None

    def _get_template_sidecar_path(self, template_id: str) -> Optional[Path]:
        for ext in ['.yaml', '.yml']:
            sidecar_path = self.templates_dir / f"{template_id}{ext}"
            if sidecar_path.exists():
                return sidecar_path
        return None
    
    def create_template_structure(self, template_id: str, description: str = ""):
        """Crée la structure pour un nouveau template"""
        # Crée le dossier masters s'il n'existe pas
        self.masters_dir.mkdir(exist_ok=True)
        
        # Ajoute au registry
        if template_id not in self.registry["templates"]:
            self.registry["templates"][template_id] = {
                "description": description,
                "base_score": 50
            }
            self._save_registry()
        
        # Crée les fichiers vides
        excel_path = self.masters_dir / f"{template_id}.xlsx"
        sidecar_path = self.templates_dir / f"{template_id}.yml"
        
        if not excel_path.exists():
            # Crée un fichier Excel vide (à compléter manuellement)
            self._create_empty_excel_template(excel_path, template_id, description)
        
        if not sidecar_path.exists():
            # Crée un sidecar de base
            self._create_base_sidecar(sidecar_path, template_id, description)
    
    def _create_empty_excel_template(self, excel_path: Path, template_id: str, description: str):
        """Crée un template Excel vide"""
        try:
            import openpyxl
            import datetime          # <-- ajouter ça ici

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Info"

            # Ajoute des informations de base
            ws['A1'] = f"Template: {template_id}"
            ws['A2'] = f"Description: {description}"
            ws['A3'] = f"Created: {datetime.datetime.now().strftime('%Y-%m-%d')}"
            ws['A5'] = "=== STRUCTURE ATTENDUE ==="
            ws['A6'] = "Créez les feuilles et tables décrites dans le sidecar YAML"

            wb.save(excel_path)
            logger.info(f"Empty Excel template created: {excel_path}")

        except ImportError:
            logger.warning("openpyxl not available, cannot create Excel template")
            excel_path.touch()

    
    def _create_base_sidecar(self, sidecar_path: Path, template_id: str, description: str):
        """Crée un sidecar YAML de base"""
        base_sidecar = {
            "id": template_id,
            "version": "1.0.0",
            "description": description,
            "sheets": {
                "Verslag": {
                    "description": "Feuille de rapport principale",
                    "required": True,
                    "fields": {
                        "vendor": {
                            "name": "AI_vendor",
                            "type": "string",
                            "required": True,
                            "description": "Fabricant du système"
                        },
                        "test_date": {
                            "name": "AI_test_date", 
                            "type": "date",
                            "required": True,
                            "description": "Date du test"
                        }
                    }
                }
            }
        }
        
        with open(sidecar_path, 'w', encoding='utf-8') as f:
            yaml.dump(base_sidecar, f, default_flow_style=False, indent=2)
        
        logger.info(f"Base sidecar created: {sidecar_path}")
    
    def _save_registry(self):
        """Sauvegarde le registry"""
        registry_path = self.templates_dir / "index.yml"
        try:
            with open(registry_path, 'w', encoding='utf-8') as f:
                yaml.dump(self.registry, f, default_flow_style=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save registry: {e}")
